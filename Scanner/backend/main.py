import io
import os
import traceback
from typing import List, Optional, Union

import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, field_validator, ConfigDict
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker

from routes.auth_routes import auth_router
from routes.supervisor_routes import supervisor_bp

# ruta base de datos
DB_PATH = os.path.join(os.path.dirname(__file__), "instance", "scanner.db")
DATABASE_URL = f"sqlite:///{DB_PATH}"

engine = create_engine(
    DATABASE_URL, connect_args={"timeout": 30.0}, future=True
)

@event.listens_for(engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    cursor = dbapi_connection.cursor()
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.execute("PRAGMA synchronous=NORMAL")
    cursor.close()

SessionLocal = sessionmaker(
    autocommit=False, autoflush=False, bind=engine, future=True
)

# ruta plantillas
PLANTILLA_DIR = os.path.join("storage", "plantilla")
PATH_PLANTILLA_PDF = os.path.join(PLANTILLA_DIR, "maestra.pdf")
PATH_PLANTILLA_XLSX = os.path.join(PLANTILLA_DIR, "maestra.xlsx")

# escalas pdf
IMG_W, IMG_H = 2551.0, 3301.0
PDF_W, PDF_H = 612.0, 792.0

app = FastAPI(title="Encuestas API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# registrar rutas
app.include_router(auth_router)
app.include_router(supervisor_bp)


# esquema rellenar encuestas
class EncuestaSchema(BaseModel):
    model_config = ConfigDict(extra="allow")
    
    nombre_empresa: str
    rut_empresa: str
    nombre_encuestado: str
    cargo: Optional[str] = ""
    correo: str
    telefono: Optional[str] = ""
    fecha: Optional[str] = None
    firma: Optional[str] = ""
    
    p1_1: Optional[str] = None
    p1_2: Optional[str] = None
    p1_3: Optional[str] = None
    p2_1: Optional[str] = None
    p2_2: Optional[str] = None
    p2_3: Optional[str] = None
    p3_1: Optional[str] = None
    p3_2: Optional[str] = None
    p3_3: Optional[str] = None

    red_mas_usa: Optional[Union[str, List[str]]] = None
    red_sigue: Optional[Union[str, List[str]]] = None

    correo_informativo: Optional[Union[int, str]] = 0
    observaciones: Optional[str] = ""

    @field_validator("correo_informativo", mode="before")
    @classmethod
    def format_correo_informativo(cls, v):
        if v is None or v == "" or v == "null":
            return 0
        if str(v).lower() in ["1", "true", "si", "sí"]:
            return 1
        return 0


# aux
def pixel_to_pdf_coords(x, y, w, h):
    return (x * (PDF_W / IMG_W)), ((IMG_H - (y + h)) * (PDF_H / IMG_H))

def obtener_id_casilla(pregunta: str, opcion: str) -> Optional[int]:
    opciones_eval = ["Siempre", "Generalmente", "Rara vez", "Nunca"]
    redes = ["Instagram", "TikTok", "Facebook", "LinkedIn", "Pinterest", "Ninguna"]
    mapeo = {
        "p1_1": range(9, 13), "p1_2": range(13, 17), "p1_3": range(17, 21),
        "p2_1": range(21, 25), "p2_2": range(25, 29), "p2_3": range(29, 33),
        "p3_1": range(33, 37), "p3_2": range(37, 41), "p3_3": range(41, 45),
    }
    if pregunta in mapeo and opcion in opciones_eval:
        return list(mapeo[pregunta])[opciones_eval.index(opcion)]
    if pregunta == "red_mas_usa" and opcion in redes:
        return [45, 47, 48, 49, 50, 51][redes.index(opcion)]
    if pregunta == "red_sigue" and opcion in redes:
        return [52, 53, 54, 56, 57, 58][redes.index(opcion)]
    if pregunta == "correo_informativo":
        return 59 if str(opcion) in ["1", "True", "true", "Sí", "si"] else 60
    return None

# Para llevar a Excel
def generar_excel_desde_bd(registro_id: int, salida_path: str):
    db = SessionLocal()
    try:
        dato = db.execute(text("SELECT * FROM dato_extraido WHERE id = :id"), {"id": registro_id}).fetchone()
        if not dato: return
        datos = dict(dato._mapping)

        wb = openpyxl.load_workbook(PATH_PLANTILLA_XLSX)
        ws = wb.active

        ws["B3"] = datos.get("nombre_empresa", "")
        ws["B4"] = datos.get("rut_empresa", "")
        ws["B5"] = datos.get("nombre_encuestado", "")
        ws["B6"] = datos.get("cargo", "")
        ws["B7"] = datos.get("correo", "")
        ws["H7"] = datos.get("telefono", "")
        ws["B8"] = datos.get("fecha", "")
        ws["H8"] = datos.get("firma", "")

        def marcar_columna_excel(fila, valor_opcion):
            if not valor_opcion: return
            val_lower = str(valor_opcion).strip().lower()
            col_meta = None
            if "siempre" in val_lower: col_meta = 8
            elif "generalmente" in val_lower: col_meta = 9
            elif "rara" in val_lower: col_meta = 11
            elif "nunca" in val_lower: col_meta = 12

            if col_meta: ws.cell(row=fila, column=col_meta).value = "X"

        marcar_columna_excel(11, datos.get("p1_1"))
        marcar_columna_excel(12, datos.get("p1_2"))
        marcar_columna_excel(13, datos.get("p1_3"))

        marcar_columna_excel(18, datos.get("p2_1"))
        marcar_columna_excel(19, datos.get("p2_2"))
        marcar_columna_excel(20, datos.get("p2_3"))

        marcar_columna_excel(25, datos.get("p3_1"))
        marcar_columna_excel(26, datos.get("p3_2"))
        marcar_columna_excel(27, datos.get("p3_3"))

        redes_map = {
            "instagram": 8, "tiktok": 9, "facebook": 10,
            "linkedin": 11, "pinterest": 12, "ninguna": 13,
        }

        red_mas = datos.get("red_mas_usa", "")
        if red_mas:
            for r_item in str(red_mas).split(","):
                clean_r = r_item.strip().lower()
                if clean_r in redes_map: ws.cell(row=31, column=redes_map[clean_r]).value = "X"

        red_sig = datos.get("red_sigue", "")
        if red_sig:
            for r_item in str(red_sig).split(","):
                clean_r = r_item.strip().lower()
                if clean_r in redes_map: ws.cell(row=32, column=redes_map[clean_r]).value = "X"

        correo_inf = str(datos.get("correo_informativo", "")).strip().lower()
        if correo_inf in ["1", "si", "sí", "true"]: ws.cell(row=35, column=8).value = "X"
        elif correo_inf in ["0", "no", "false"]: ws.cell(row=35, column=9).value = "X"

        ws["B39"] = datos.get("observaciones", "")
        wb.save(salida_path)
    finally:
        db.close()


# endpoints
@app.post("/api/encuesta/guardar")
@app.post("/api/encuesta")
def guardar_y_generar_encuesta(payload: EncuestaSchema):
    db = SessionLocal()
    try:
        datos = payload.model_dump()
        extra_datos = payload.model_extra or {}

        def limpiar_evaluacion(val):
            if not val: return None
            v_lower = str(val).lower()
            if "siempre" in v_lower: return "Siempre"
            if "generalmente" in v_lower: return "Generalmente"
            if "rara" in v_lower: return "Rara vez"
            if "nunca" in v_lower: return "Nunca"
            return val

        datos['p1_1'] = limpiar_evaluacion(datos.get('p1_1') or extra_datos.get('pedidos_completos'))
        datos['p1_2'] = limpiar_evaluacion(datos.get('p1_2') or extra_datos.get('pedidos_rapidos'))
        datos['p1_3'] = limpiar_evaluacion(datos.get('p1_3') or extra_datos.get('respuestas_oportunas'))
        
        datos['p2_1'] = limpiar_evaluacion(datos.get('p2_1') or extra_datos.get('producto_bien_presentado'))
        datos['p2_2'] = limpiar_evaluacion(datos.get('p2_2') or extra_datos.get('producto_buena_calidad'))
        datos['p2_3'] = limpiar_evaluacion(datos.get('p2_3') or extra_datos.get('informacion_productos_nuevos'))
        
        datos['p3_1'] = limpiar_evaluacion(datos.get('p3_1') or extra_datos.get('contacto_con_ejecutivo'))
        datos['p3_2'] = limpiar_evaluacion(datos.get('p3_2') or extra_datos.get('calidad_atencion'))
        datos['p3_3'] = limpiar_evaluacion(datos.get('p3_3') or extra_datos.get('personal_domina_informacion'))

        red_usa = datos.get('red_mas_usa') or extra_datos.get('red_social_usa', "")
        red_sigue = datos.get('red_sigue') or extra_datos.get('red_social_sigue', "")
        
        if isinstance(red_usa, list): red_usa = ", ".join(red_usa)
        if isinstance(red_sigue, list): red_sigue = ", ".join(red_sigue)
        
        datos['red_mas_usa'] = red_usa
        datos['red_sigue'] = red_sigue
        datos['observaciones'] = datos.get('observaciones') or extra_datos.get('obs_recomen')

        # Detección de id_usuario
        id_usuario_real = extra_datos.get("id_usuario") or datos.get("id_usuario")
        if not id_usuario_real:
            primer_usr = db.execute(text("SELECT id_usuario FROM usuario LIMIT 1")).fetchone()
            id_usuario_real = primer_usr[0] if primer_usr else 3

        # insertar encuesta
        enc_res = db.execute(
            text(
                "INSERT INTO encuesta (id_usuario, empresa, rut, encuestado, cargo,"
                " correo, telefono, fecha) VALUES (:id_usuario, :empresa, :rut,"
                " :encuestado, :cargo, :correo, :telefono, :fecha)"
            ),
            {
                "id_usuario": id_usuario_real,
                "empresa": datos.get("nombre_empresa", ""),
                "rut": datos.get("rut_empresa", ""),
                "encuestado": datos.get("nombre_encuestado", ""),
                "cargo": datos.get("cargo", ""),
                "correo": datos.get("correo", ""),
                "telefono": datos.get("telefono", ""),
                "fecha": datos.get("fecha", ""),
            },
        )
        db.commit()
        registro_id = enc_res.lastrowid

        # insertar dato_extraido acorde al id
        db.execute(
            text("""
                INSERT INTO dato_extraido (
                    id, nombre_empresa, rut_empresa, nombre_encuestado, cargo, correo, telefono, fecha, firma,
                    p1_1, p1_2, p1_3, p2_1, p2_2, p2_3, p3_1, p3_2, p3_3, 
                    red_mas_usa, red_sigue, correo_informativo, observaciones
                ) VALUES (
                    :id, :nombre_empresa, :rut_empresa, :nombre_encuestado, :cargo, :correo, :telefono, :fecha, :firma,
                    :p1_1, :p1_2, :p1_3, :p2_1, :p2_2, :p2_3, :p3_1, :p3_2, :p3_3, 
                    :red_mas_usa, :red_sigue, :correo_informativo, :observaciones
                )
            """),
            {**datos, "id": registro_id},
        )
        db.commit()

        rut_clean = str(datos.get("rut_empresa", "000")).replace(".", "").replace("-", "")
        pdf_filename = f"encuesta_{registro_id}_{rut_clean}.pdf"
        excel_filename = f"encuesta_{registro_id}_{rut_clean}.xlsx"

        ruta_pdf = os.path.join("storage", "pdf_generado", pdf_filename)
        ruta_excel = os.path.join("storage", "pdf_generado", excel_filename)

        os.makedirs(os.path.join("storage", "pdf_generado"), exist_ok=True)
        generar_excel_desde_bd(registro_id, ruta_excel)

        # asignar documento según id
        db.execute(
            text("INSERT INTO DOCUMENTO (id_plantilla, id_vendedor, id_estado, ruta_pdf_final) VALUES (1, :vendedor, 1, :pdf)"),
            {"vendedor": id_usuario_real, "pdf": ruta_pdf},
        )
        db.commit()

        return {
            "status": "success",
            "id_encuesta": registro_id,
            "mensaje": "encuesta guardada exitosamente",
            "archivos": {"pdf": ruta_pdf, "excel": ruta_excel},
        }
    except Exception as e:
        db.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()
        
@app.get("/api/encuesta/detalle/{registro_id}")
def obtener_detalle_encuesta(registro_id: int):
    db = SessionLocal()
    try:
        enc = db.execute(text("SELECT * FROM encuesta WHERE id_encuesta = :id"), {"id": registro_id}).mappings().fetchone()
        if not enc:
            raise HTTPException(status_code=404, detail="Encuesta no encontrada")
            
        det = db.execute(text("SELECT * FROM dato_extraido WHERE id = :id"), {"id": registro_id}).mappings().fetchone()
        
        if not det:
            det = db.execute(
                text("SELECT * FROM dato_extraido WHERE rut_empresa = :rut AND fecha = :fecha ORDER BY id DESC LIMIT 1"),
                {"rut": enc["rut"], "fecha": enc["fecha"]}
            ).mappings().fetchone()
            
        resultado = dict(enc)
        if det:
            det_dict = dict(det)
            resultado.update(det_dict)
            
            # traductor pdf de react
            resultado['pedidos_completos'] = det_dict.get('p1_1')
            resultado['pedidos_rapidos'] = det_dict.get('p1_2')
            resultado['respuestas_oportunas'] = det_dict.get('p1_3')
            
            resultado['producto_bien_presentado'] = det_dict.get('p2_1')
            resultado['producto_buena_calidad'] = det_dict.get('p2_2')
            resultado['informacion_productos_nuevos'] = det_dict.get('p2_3')
            
            resultado['contacto_con_ejecutivo'] = det_dict.get('p3_1')
            resultado['calidad_atencion'] = det_dict.get('p3_2')
            resultado['personal_domina_informacion'] = det_dict.get('p3_3')
            
            if det_dict.get('red_mas_usa'):
                resultado['red_social_usa'] = [r.strip() for r in str(det_dict.get('red_mas_usa')).split(',')]
            if det_dict.get('red_sigue'):
                resultado['red_social_sigue'] = [r.strip() for r in str(det_dict.get('red_sigue')).split(',')]
            
        # Obtener nombre del usuario
        resultado['usuario'] = 'Vendedor'
        try:
            if 'id_usuario' in resultado and resultado['id_usuario']:
                usr = db.execute(text("SELECT * FROM usuario WHERE id_usuario = :uid OR id = :uid"), {"uid": resultado['id_usuario']}).mappings().fetchone()
                if usr:
                    nombre = usr.get('nombre', '')
                    apellido = usr.get('apellido', '')
                    if nombre or apellido:
                        resultado['usuario'] = f"{nombre} {apellido}".strip()
                    else:
                        for key in ['username', 'usuario', 'mail', 'user']:
                            if key in usr and usr[key]:
                                resultado['usuario'] = usr[key]
                                break
        except Exception:
            pass

        return resultado
    finally:
        db.close()


@app.get("/api/encuesta/listar")
def listar_encuestas():
    db = SessionLocal()
    try:
        # ¡ESTA ES LA CONSULTA CORREGIDA! 
        # Ahora FastAPI devolverá los nombres igual que la versión de Flask
        query = text("""
            SELECT 
                e.*, 
                (SELECT user FROM USUARIO WHERE id_usuario = e.id_usuario) AS username,
                (SELECT nombre FROM USUARIO WHERE id_usuario = e.id_usuario) AS nombre,
                (SELECT apellido FROM USUARIO WHERE id_usuario = e.id_usuario) AS apellido
            FROM encuesta e 
            ORDER BY e.id_encuesta DESC
        """)
        resultados = db.execute(query).mappings().fetchall()
        return [dict(row) for row in resultados]
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.get("/api/encuesta/exportar/{tipo}/{registro_id}")
def exportar_encuesta(tipo: str, registro_id: int):
    db = SessionLocal()
    try:
        enc = db.execute(
            text("SELECT rut FROM encuesta WHERE id_encuesta = :id"), 
            {"id": registro_id}
        ).fetchone()
        
        if not enc:
            raise HTTPException(status_code=404, detail="Encuesta no encontrada")
        
        rut_clean = str(enc[0] or "000").replace(".", "").replace("-", "")
        
        if tipo == "pdf":
            filename = f"encuesta_{registro_id}_{rut_clean}.pdf"
        elif tipo == "excel":
            filename = f"encuesta_{registro_id}_{rut_clean}.xlsx"
        else:
            raise HTTPException(status_code=400, detail="Tipo de archivo inválido")
            
        filepath = os.path.join("storage", "pdf_generado", filename)
        
        if os.path.exists(filepath):
            return FileResponse(filepath, filename=filename)
        else:
            raise HTTPException(status_code=404, detail="El archivo físico no existe en el servidor")
    finally:
        db.close()


# endpoint para actualizar
@app.put("/api/encuesta/actualizar/{registro_id}")
def actualizar_encuesta(registro_id: int, payload: dict):
    db = SessionLocal()
    try:
        # Actualziar cabecera de la encuesta
        db.execute(
            text("""
                UPDATE encuesta 
                SET empresa = :empresa, rut = :rut, encuestado = :encuestado, 
                    cargo = :cargo, correo = :correo, telefono = :telefono, fecha = :fecha
                WHERE id_encuesta = :id
            """),
            {
                "id": registro_id,
                "empresa": payload.get("empresa", payload.get("nombre_empresa", "")),
                "rut": payload.get("rut", payload.get("rut_empresa", "")),
                "encuestado": payload.get("encuestado", payload.get("nombre_encuestado", "")),
                "cargo": payload.get("cargo", ""),
                "correo": payload.get("correo", ""),
                "telefono": payload.get("telefono", ""),
                "fecha": payload.get("fecha", "")
            }
        )

        def limpiar_evaluacion(val):
            if not val: return None
            v_lower = str(val).lower()
            if "siempre" in v_lower: return "Siempre"
            if "generalmente" in v_lower: return "Generalmente"
            if "rara" in v_lower: return "Rara vez"
            if "nunca" in v_lower: return "Nunca"
            return val

        red_usa = payload.get("red_mas_usa", payload.get("red_social_usa", ""))
        if isinstance(red_usa, list): red_usa = ", ".join(red_usa)
        
        red_sigue = payload.get("red_sigue", payload.get("red_social_sigue", ""))
        if isinstance(red_sigue, list): red_sigue = ", ".join(red_sigue)

        params_det = {
            "id": registro_id,
            "empresa": payload.get("empresa", payload.get("nombre_empresa", "")),
            "rut": payload.get("rut", payload.get("rut_empresa", "")),
            "encuestado": payload.get("encuestado", payload.get("nombre_encuestado", "")),
            "cargo": payload.get("cargo", ""),
            "correo": payload.get("correo", ""),
            "telefono": payload.get("telefono", ""),
            "fecha": payload.get("fecha", ""),
            "p1_1": limpiar_evaluacion(payload.get("p1_1") or payload.get("pedidos_completos")),
            "p1_2": limpiar_evaluacion(payload.get("p1_2") or payload.get("pedidos_rapidos")),
            "p1_3": limpiar_evaluacion(payload.get("p1_3") or payload.get("respuestas_oportunas")),
            "p2_1": limpiar_evaluacion(payload.get("p2_1") or payload.get("producto_bien_presentado")),
            "p2_2": limpiar_evaluacion(payload.get("p2_2") or payload.get("producto_buena_calidad")),
            "p2_3": limpiar_evaluacion(payload.get("p2_3") or payload.get("informacion_productos_nuevos")),
            "p3_1": limpiar_evaluacion(payload.get("p3_1") or payload.get("contacto_con_ejecutivo")),
            "p3_2": limpiar_evaluacion(payload.get("p3_2") or payload.get("calidad_atencion")),
            "p3_3": limpiar_evaluacion(payload.get("p3_3") or payload.get("personal_domina_informacion")),
            "red_mas_usa": red_usa,
            "red_sigue": red_sigue,
            "correo_informativo": payload.get("correo_informativo", 0),
            "observaciones": payload.get("observaciones", payload.get("obs_recomen", ""))
        }

        # Verificar si existe en dato_extraido
        existe_det = db.execute(text("SELECT id FROM dato_extraido WHERE id = :id"), {"id": registro_id}).fetchone()

        if existe_det:
            # actualizar si ya existe
            db.execute(
                text("""
                    UPDATE dato_extraido 
                    SET nombre_empresa = :empresa, rut_empresa = :rut, nombre_encuestado = :encuestado,
                        cargo = :cargo, correo = :correo, telefono = :telefono, fecha = :fecha,
                        p1_1 = :p1_1, p1_2 = :p1_2, p1_3 = :p1_3,
                        p2_1 = :p2_1, p2_2 = :p2_2, p2_3 = :p2_3,
                        p3_1 = :p3_1, p3_2 = :p3_2, p3_3 = :p3_3,
                        red_mas_usa = :red_mas_usa, red_sigue = :red_sigue,
                        correo_informativo = :correo_informativo,
                        observaciones = :observaciones
                    WHERE id = :id
                """),
                params_det
            )
        else:
            #detalle asociado
            db.execute(
                text("""
                    INSERT INTO dato_extraido (
                        id, nombre_empresa, rut_empresa, nombre_encuestado, cargo, correo, telefono, fecha, firma,
                        p1_1, p1_2, p1_3, p2_1, p2_2, p2_3, p3_1, p3_2, p3_3, 
                        red_mas_usa, red_sigue, correo_informativo, observaciones
                    ) VALUES (
                        :id, :empresa, :rut, :encuestado, :cargo, :correo, :telefono, :fecha, '',
                        :p1_1, :p1_2, :p1_3, :p2_1, :p2_2, :p2_3, :p3_1, :p3_2, :p3_3, 
                        :red_mas_usa, :red_sigue, :correo_informativo, :observaciones
                    )
                """),
                params_det
            )

        db.commit()
        return {"status": "success", "mensaje": "Encuesta actualizada correctamente"}
    except Exception as e:
        db.rollback()
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@app.delete("/api/encuesta/eliminar/{registro_id}")
def eliminar_encuesta(registro_id: int):
    db = SessionLocal()
    try:
        db.execute(text("DELETE FROM encuesta WHERE id_encuesta = :id"), {"id": registro_id})
        db.execute(text("DELETE FROM dato_extraido WHERE id = :id"), {"id": registro_id})
        db.commit()
        return {"status": "success", "mensaje": "Encuesta eliminada"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()        

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8082, reload=True)