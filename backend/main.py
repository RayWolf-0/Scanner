import io
import os
from typing import List, Optional, Union
import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, field_validator
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker

# --- CONFIGURACIÓN DE BASE DE DATOS (Ruta exacta a instance/scanner.db) ---
DB_PATH = os.path.join(os.path.dirname(__file__), "instance", "scanner.db")
DATABASE_URL = f"sqlite:///{DB_PATH}"

engine = create_engine(
    DATABASE_URL, connect_args={"timeout": 30.0}, future=True
)


@event.listens_for(engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
  cursor = dbapi_connection.cursor()
  cursor.execute("PRAGMA journal_mode=WAL")
  cursor.execute("PRAGMA synchronous=NORMAL")
  cursor.close()


SessionLocal = sessionmaker(
    autocommit=False, autoflush=False, bind=engine, future=True
)

# --- RUTAS DE PLANTILLAS ---
PLANTILLA_DIR = os.path.join("storage", "plantilla")
PATH_PLANTILLA_PDF = os.path.join(PLANTILLA_DIR, "maestra.pdf")
PATH_PLANTILLA_XLSX = os.path.join(PLANTILLA_DIR, "maestra.xlsx")

# --- CONSTANTES DE ESCALA PDF ---
IMG_W, IMG_H = 2551.0, 3301.0
PDF_W, PDF_H = 612.0, 792.0

app = FastAPI(title="Encuestas API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- ESQUEMA DE ENTRADA DESDE REACT (RellenarEncuesta.jsx) ---
# --- ESQUEMA DE ENTRADA DESDE REACT (RellenarEncuesta.jsx) ---
class EncuestaSchema(BaseModel):
  nombre_empresa: str
  rut_empresa: str
  nombre_encuestado: str
  cargo: Optional[str] = ""
  correo: str
  telefono: Optional[str] = ""
  fecha: Optional[str] = None
  firma: Optional[str] = ""
  p1_1: Optional[str] = None
  p1_2: Optional[str] = None
  p1_3: Optional[str] = None
  p2_1: Optional[str] = None
  p2_2: Optional[str] = None
  p2_3: Optional[str] = None
  p3_1: Optional[str] = None
  p3_2: Optional[str] = None
  p3_3: Optional[str] = None

  red_mas_usa: Optional[Union[str, List[str]]] = None
  red_sigue: Optional[Union[str, List[str]]] = None

  correo_informativo: Optional[Union[int, str]] = 0
  observaciones: Optional[str] = ""

  @field_validator("red_mas_usa", "red_sigue", mode="before")
  @classmethod
  def format_list_to_string(cls, v):
    if isinstance(v, list):
      return ", ".join(str(item) for item in v)
    return v

  @field_validator("correo_informativo", mode="before")
  @classmethod
  def format_correo_informativo(cls, v):
    # Si viene vacío, nulo o texto inválido, lo convertimos en 0 por defecto
    if v is None or v == "" or v == "null":
      return 0
    # Si viene como texto o booleano afirmativo, lo pasamos a 1, de lo contrario 0
    if str(v).lower() in ["1", "true", "si", "sí"]:
      return 1
    return 0


# --- AUXILIARES ---
def pixel_to_pdf_coords(x, y, w, h):
  return (x * (PDF_W / IMG_W)), ((IMG_H - (y + h)) * (PDF_H / IMG_H))


def obtener_id_casilla(pregunta: str, opcion: str) -> Optional[int]:
  opciones_eval = ["Siempre", "Generalmente", "Rara vez", "Nunca"]
  redes = ["Instagram", "TikTok", "Facebook", "LinkedIn", "Pinterest", "Ninguna"]
  mapeo = {
      "p1_1": range(9, 13),
      "p1_2": range(13, 17),
      "p1_3": range(17, 21),
      "p2_1": range(21, 25),
      "p2_2": range(25, 29),
      "p2_3": range(29, 33),
      "p3_1": range(33, 37),
      "p3_2": range(37, 41),
      "p3_3": range(41, 45),
  }
  if pregunta in mapeo and opcion in opciones_eval:
    return list(mapeo[pregunta])[opciones_eval.index(opcion)]
  if pregunta == "red_mas_usa" and opcion in redes:
    return [45, 47, 48, 49, 50, 51][redes.index(opcion)]
  if pregunta == "red_sigue" and opcion in redes:
    return [52, 53, 54, 56, 57, 58][redes.index(opcion)]
  if pregunta == "correo_informativo":
    return 59 if str(opcion) in ["1", "True", "true", "Sí", "si"] else 60
  return None


# --- LÓGICA DE PDF DESDE BD ---
def generar_pdf_desde_bd(registro_id: int, salida_path: str):
  db = SessionLocal()
  try:
    dato = db.execute(
        text("SELECT * FROM dato_extraido WHERE id = :id"), {"id": registro_id}
    ).fetchone()
    if not dato:
      return
    datos = dict(dato._mapping)

    # Uso estricto de CAMPO_PLANTILLA y sus nombres reales de columnas
    mapa_filas = db.execute(
        text(
            "SELECT id_campo, nombre_campo, tipo_dato, pos_x, pos_y, ancho, alto"
            " FROM CAMPO_PLANTILLA"
        )
    ).fetchall()
    mapa = {
        row.id_campo: {
            "nombre": row.nombre_campo,
            "tipo": row.tipo_dato,
            "x": row.pos_x,
            "y": row.pos_y,
            "w": row.ancho,
            "h": row.alto,
        }
        for row in mapa_filas
    }

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFont("Helvetica-Bold", 10)

    text_fields = {
        1: datos.get("nombre_empresa"),
        2: datos.get("rut_empresa"),
        3: datos.get("nombre_encuestado"),
        4: datos.get("cargo"),
        5: datos.get("correo"),
        6: datos.get("telefono"),
        7: datos.get("fecha"),
        8: datos.get("firma"),
        61: datos.get("observaciones"),
    }
    for fid, val in text_fields.items():
      if val and fid in mapa:
        px, py = pixel_to_pdf_coords(
            mapa[fid]["x"], mapa[fid]["y"], mapa[fid]["w"], mapa[fid]["h"]
        )
        can.drawString(px, py, str(val))

    def marcar_casillas(nombre_campo_bd):
      valor = datos.get(nombre_campo_bd)
      if not valor:
        return []
      opciones = [o.strip() for o in str(valor).split(",")]
      ids_a_marcar = []
      for op in opciones:
        cid = obtener_id_casilla(nombre_campo_bd, op)
        if cid:
          ids_a_marcar.append(cid)
      return ids_a_marcar

    casillas_a_marcar = []
    for p in [
        "p1_1",
        "p1_2",
        "p1_3",
        "p2_1",
        "p2_2",
        "p2_3",
        "p3_1",
        "p3_2",
        "p3_3",
    ]:
      casillas_a_marcar.extend(marcar_casillas(p))

    casillas_a_marcar.extend(marcar_casillas("red_mas_usa"))
    casillas_a_marcar.extend(marcar_casillas("red_sigue"))
    casillas_a_marcar.extend(marcar_casillas("correo_informativo"))

    for cid in [c for c in casillas_a_marcar if c and c in mapa]:
      px, py = pixel_to_pdf_coords(
          mapa[cid]["x"], mapa[cid]["y"], mapa[cid]["w"], mapa[cid]["h"]
      )
      can.drawString(px, py, "X")

    can.save()
    packet.seek(0)
    output_writer = PdfWriter()

    template_reader = PdfReader(PATH_PLANTILLA_PDF)
    overlay_reader = PdfReader(packet)

    for index, page in enumerate(template_reader.pages):
      if index == 0 and len(overlay_reader.pages) > 0:
        page.merge_page(overlay_reader.pages[0])
      output_writer.add_page(page)

    with open(salida_path, "wb") as f:
      output_writer.write(f)
  finally:
    db.close()


# --- LÓGICA DE EXCEL DESDE BD ---
def generar_excel_desde_bd(registro_id: int, salida_path: str):
  db = SessionLocal()
  try:
    dato = db.execute(
        text("SELECT * FROM dato_extraido WHERE id = :id"), {"id": registro_id}
    ).fetchone()
    if not dato:
      return
    datos = dict(dato._mapping)

    wb = openpyxl.load_workbook(PATH_PLANTILLA_XLSX)
    ws = wb.active

    ws["B3"] = datos.get("nombre_empresa", "")
    ws["B4"] = datos.get("rut_empresa", "")
    ws["B5"] = datos.get("nombre_encuestado", "")
    ws["B6"] = datos.get("cargo", "")
    ws["B7"] = datos.get("correo", "")
    ws["H7"] = datos.get("telefono", "")
    ws["B8"] = datos.get("fecha", "")
    ws["H8"] = datos.get("firma", "")

    def marcar_columna_excel(fila, valor_opcion):
      if not valor_opcion:
        return
      val_lower = str(valor_opcion).strip().lower()
      col_meta = None
      if "siempre" in val_lower:
        col_meta = 8
      elif "generalmente" in val_lower:
        col_meta = 9
      elif "rara" in val_lower:
        col_meta = 11
      elif "nunca" in val_lower:
        col_meta = 12

      if col_meta:
        ws.cell(row=fila, column=col_meta).value = "X"

    marcar_columna_excel(11, datos.get("p1_1"))
    marcar_columna_excel(12, datos.get("p1_2"))
    marcar_columna_excel(13, datos.get("p1_3"))

    marcar_columna_excel(18, datos.get("p2_1"))
    marcar_columna_excel(19, datos.get("p2_2"))
    marcar_columna_excel(20, datos.get("p2_3"))

    marcar_columna_excel(25, datos.get("p3_1"))
    marcar_columna_excel(26, datos.get("p3_2"))
    marcar_columna_excel(27, datos.get("p3_3"))

    redes_map = {
        "instagram": 8,
        "tiktok": 9,
        "facebook": 10,
        "linkedin": 11,
        "pinterest": 12,
        "ninguna": 13,
    }

    red_mas = datos.get("red_mas_usa", "")
    if red_mas:
      for r_item in str(red_mas).split(","):
        clean_r = r_item.strip().lower()
        if clean_r in redes_map:
          ws.cell(row=31, column=redes_map[clean_r]).value = "X"

    red_sig = datos.get("red_sigue", "")
    if red_sig:
      for r_item in str(red_sig).split(","):
        clean_r = r_item.strip().lower()
        if clean_r in redes_map:
          ws.cell(row=32, column=redes_map[clean_r]).value = "X"

    correo_inf = str(datos.get("correo_informativo", "")).strip().lower()
    if correo_inf in ["1", "si", "sí", "true"]:
      ws.cell(row=35, column=8).value = "X"
    elif correo_inf in ["0", "no", "false"]:
      ws.cell(row=35, column=9).value = "X"

    ws["B39"] = datos.get("observaciones", "")

    wb.save(salida_path)
  finally:
    db.close()


# --- ENDPOINTS ---
@app.post("/api/encuesta/guardar")
@app.post("/api/encuesta")
def guardar_y_generar_encuesta(payload: EncuestaSchema):
  db = SessionLocal()
  try:
    datos = payload.model_dump()

    # 1. Inserción en tabla encuesta
    db.execute(
        text(
            "INSERT INTO encuesta (id_usuario, empresa, rut, encuestado, cargo,"
            " correo, telefono, fecha) VALUES (:id_usuario, :empresa, :rut,"
            " :encuestado, :cargo, :correo, :telefono, :fecha)"
        ),
        {
            "id_usuario": 1,
            "empresa": datos.get("nombre_empresa", ""),
            "rut": datos.get("rut_empresa", ""),
            "encuestado": datos.get("nombre_encuestado", ""),
            "cargo": datos.get("cargo", ""),
            "correo": datos.get("correo", ""),
            "telefono": datos.get("telefono", ""),
            "fecha": datos.get("fecha", ""),
        },
    )
    db.commit()

    # 2. Inserción en tabla dato_extraido
    dato_res = db.execute(
        text("""
            INSERT INTO dato_extraido (
                nombre_empresa, rut_empresa, nombre_encuestado, cargo, correo, telefono, fecha, firma,
                p1_1, p1_2, p1_3, p2_1, p2_2, p2_3, p3_1, p3_2, p3_3, 
                red_mas_usa, red_sigue, correo_informativo, observaciones
            ) VALUES (
                :nombre_empresa, :rut_empresa, :nombre_encuestado, :cargo, :correo, :telefono, :fecha, :firma,
                :p1_1, :p1_2, :p1_3, :p2_1, :p2_2, :p2_3, :p3_1, :p3_2, :p3_3, 
                :red_mas_usa, :red_sigue, :correo_informativo, :observaciones
            )
        """),
        datos,
    )
    db.commit()
    registro_id = dato_res.lastrowid

    rut_clean = str(datos.get("rut_empresa", "000")).replace(".", "").replace("-", "")
    pdf_filename = f"encuesta_{registro_id}_{rut_clean}.pdf"
    excel_filename = f"encuesta_{registro_id}_{rut_clean}.xlsx"

    generar_pdf_desde_bd(registro_id, pdf_filename)
    generar_excel_desde_bd(registro_id, excel_filename)

    # 3. Inserción en DOCUMENTO (respetando sus columnas reales)
    db.execute(
        text("INSERT INTO DOCUMENTO (ruta_pdf_final) VALUES (:pdf)"),
        {"pdf": pdf_filename},
    )
    db.commit()

    return {
        "ok": True,
        "id_registro": registro_id,
        "mensaje": "Encuesta guardada y documentos rellenados con éxito",
        "archivos": {"pdf": pdf_filename, "excel": excel_filename},
    }
  except Exception as e:
    db.rollback()
    import traceback

    traceback.print_exc()
    raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/descargar/{filename}")
def descargar_archivo(filename: str):
  if os.path.exists(filename):
    return FileResponse(filename)
  raise HTTPException(status_code=404, detail="Archivo no encontrado")


if __name__ == "__main__":
  import uvicorn

  uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)